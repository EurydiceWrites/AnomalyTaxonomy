#!/usr/bin/env python3
"""
Phase 2b QA Triage Generator for the Mack-Bullard UFO Matrix.

Aligns AI blind-coded output against Bullard's ground truth from ufo_matrix.db,
performs draft divergence triage, and produces a formatted .xlsx for human review.

Usage:
    python3 qa_triage.py \
        --csv phase2_results.csv \
        --db ufo_matrix.db \
        --case 192g \
        --profile baseline_test \
        --output qa_triage_output.xlsx \
        [--prev-match 99 --prev-miss 23 --prev-extra 50]
"""

import argparse
import csv
import json
import sqlite3
import sys
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ── Data Loading ──

def load_ai_results(input_path):
    """Load phase2 results from CSV or JSON, auto-detected by file extension."""
    rows = []
    ext = Path(input_path).suffix.lower()

    if ext == '.json':
        with open(input_path, 'r', encoding='utf-8-sig') as f:
            data = json.load(f)
        events = data.get('ai_events', data) if isinstance(data, dict) else data
        for ev in events:
            rows.append({
                'seq': int(ev.get('sequence', 0)),
                'chunk': int(ev.get('chunk', 0)),
                'code': ev.get('motif_code', '').strip(),
                'citation': ev.get('citation', '').strip(),
                'reasoning': ev.get('reasoning', '').strip(),
                'voice_speaker': ev.get('voice_speaker', '').strip(),
                'voice_content_type': ev.get('voice_content_type', '').strip(),
            })
    else:
        # Default: CSV
        with open(input_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append({
                    'seq': int(row['sequence']),
                    'chunk': int(row['chunk']),
                    'code': row['motif_code'].strip(),
                    'citation': row['citation'].strip(),
                    'reasoning': row['reasoning'].strip(),
                    'voice_speaker': row.get('voice_speaker', '').strip(),
                    'voice_content_type': row.get('voice_content_type', '').strip(),
                })
    return rows


def load_ground_truth(db_path, case_number):
    """Load ground truth events for a case from ufo_matrix.db."""
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    # Get encounter ID
    cur.execute("SELECT Encounter_ID FROM Encounters WHERE Case_Number = ?", (case_number,))
    result = cur.fetchone()
    if not result:
        print(f"ERROR: Case '{case_number}' not found in Encounters table.", file=sys.stderr)
        conn.close()
        sys.exit(1)
    encounter_id = result[0]

    # Get events with motif descriptions
    cur.execute("""
        SELECT ee.Sequence_Order, ee.Motif_Code, md.motif_description, ee.Source_Citation
        FROM Encounter_Events ee
        LEFT JOIN Motif_Dictionary md ON ee.Motif_Code = md.Motif_Code
        WHERE ee.Encounter_ID = ?
        ORDER BY ee.Sequence_Order
    """, (encounter_id,))

    gt_rows = [{
        'seq': r[0], 'code': r[1], 'desc': r[2] or '', 'citation': r[3] or ''
    } for r in cur.fetchall()]

    conn.close()
    return gt_rows


def load_motif_dict(db_path):
    """Load the full motif dictionary for description lookups."""
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("SELECT Motif_Code, motif_description FROM Motif_Dictionary")
    d = {r[0]: r[1] for r in cur.fetchall()}
    conn.close()
    return d


# ── Text Overlap Utilities ──

STOPWORDS = {
    'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for',
    'of', 'with', 'by', 'from', 'as', 'is', 'was', 'are', 'were', 'be',
    'been', 'being', 'have', 'has', 'had', 'do', 'does', 'did', 'will',
    'would', 'could', 'should', 'may', 'might', 'shall', 'can', 'not',
    'no', 'nor', 'so', 'yet', 'both', 'each', 'her', 'his', 'its', 'my',
    'our', 'their', 'your', 'he', 'she', 'it', 'we', 'they', 'them',
    'him', 'this', 'that', 'these', 'those', 'then', 'than', 'into',
    'up', 'out', 'about', 'if', 'when', 'where', 'how', 'all', 'very',
}


def content_words(text):
    """Extract content words from text: lowercase, > 2 chars, not stopwords."""
    if not text:
        return set()
    words = set()
    for w in text.lower().split():
        cleaned = ''.join(c for c in w if c.isalnum())
        if len(cleaned) > 2 and cleaned not in STOPWORDS:
            words.add(cleaned)
    return words


def text_overlap_score(text_a, text_b):
    """Count shared content words between two text strings."""
    return len(content_words(text_a) & content_words(text_b))


# ── Alignment ──

def align(gt_rows, ai_rows, motif_dict):
    """
    Hybrid citation-aware alignment of GT codes against AI output.

    For codes that appear only once: greedy match (unchanged behavior).
    For codes that appear multiple times: score candidates by text overlap
    (primary) with sequence proximity as tiebreaker.

    Returns a list of dicts, each with 'type' (MATCH/MISS/AI EXTRA)
    and all relevant fields from both sides.
    """
    matched_ai_indices = set()
    output = []

    gt_total = len(gt_rows)
    ai_total = len(ai_rows)

    for gt_idx, gt in enumerate(gt_rows):
        # Find all unmatched AI rows with the same code
        candidates = []
        for ai_idx, ai in enumerate(ai_rows):
            if ai_idx in matched_ai_indices:
                continue
            if ai['code'] == gt['code']:
                candidates.append(ai_idx)

        if len(candidates) == 0:
            # No match — MISS
            output.append({
                'type': 'MISS',
                'gt_seq': gt['seq'],
                'bullard_code': gt['code'],
                'bullard_desc': gt['desc'],
                'bullard_citation': gt['citation'],
                'ai_code': '', 'ai_desc': '', 'ai_citation': '',
                'ai_reasoning': '', 'voice_speaker': '', 'voice_content_type': '',
                'ai_seq': '', 'ai_chunk': '',
            })
        elif len(candidates) == 1:
            # Single match — pair directly (unchanged behavior)
            ai_idx = candidates[0]
            ai = ai_rows[ai_idx]
            output.append({
                'type': 'MATCH',
                'gt_seq': gt['seq'],
                'bullard_code': gt['code'],
                'bullard_desc': gt['desc'],
                'bullard_citation': gt['citation'],
                'ai_code': ai['code'],
                'ai_desc': motif_dict.get(ai['code'], ''),
                'ai_citation': ai['citation'],
                'ai_reasoning': ai['reasoning'],
                'voice_speaker': ai.get('voice_speaker', ''),
                'voice_content_type': ai.get('voice_content_type', ''),
                'ai_seq': ai['seq'],
                'ai_chunk': ai['chunk'],
            })
            matched_ai_indices.add(ai_idx)
        else:
            # Multiple candidates — score by text overlap + sequence proximity
            best_idx = None
            best_score = (-1, float('inf'))  # (overlap DESC, proximity ASC)

            gt_position = gt_idx / gt_total if gt_total > 0 else 0

            for ai_idx in candidates:
                ai = ai_rows[ai_idx]
                overlap = text_overlap_score(gt['citation'], ai['citation'])
                ai_position = ai_idx / ai_total if ai_total > 0 else 0
                proximity = abs(gt_position - ai_position)
                score = (overlap, -proximity)  # higher overlap better, lower proximity better

                if score > best_score:
                    best_score = score
                    best_idx = ai_idx

            ai = ai_rows[best_idx]
            output.append({
                'type': 'MATCH',
                'gt_seq': gt['seq'],
                'bullard_code': gt['code'],
                'bullard_desc': gt['desc'],
                'bullard_citation': gt['citation'],
                'ai_code': ai['code'],
                'ai_desc': motif_dict.get(ai['code'], ''),
                'ai_citation': ai['citation'],
                'ai_reasoning': ai['reasoning'],
                'voice_speaker': ai.get('voice_speaker', ''),
                'voice_content_type': ai.get('voice_content_type', ''),
                'ai_seq': ai['seq'],
                'ai_chunk': ai['chunk'],
            })
            matched_ai_indices.add(best_idx)

    for ai_idx, ai in enumerate(ai_rows):
        if ai_idx not in matched_ai_indices:
            output.append({
                'type': 'AI EXTRA',
                'gt_seq': '',
                'bullard_code': '', 'bullard_desc': '', 'bullard_citation': '',
                'ai_code': ai['code'],
                'ai_desc': motif_dict.get(ai['code'], ''),
                'ai_citation': ai['citation'],
                'ai_reasoning': ai['reasoning'],
                'voice_speaker': ai.get('voice_speaker', ''),
                'voice_content_type': ai.get('voice_content_type', ''),
                'ai_seq': ai['seq'],
                'ai_chunk': ai['chunk'],
            })

    # Sort: MATCH/MISS by gt_seq, then AI EXTRA by chunk/seq
    matches_misses = sorted(
        [r for r in output if r['type'] in ('MATCH', 'MISS')],
        key=lambda r: r['gt_seq']
    )
    ai_extras = sorted(
        [r for r in output if r['type'] == 'AI EXTRA'],
        key=lambda r: (r['ai_chunk'], r['ai_seq'])
    )

    return matches_misses + ai_extras


# ── Draft Triage ──

def get_family(code):
    """Extract motif family prefix (letter + first digit), e.g. 'B2' from 'B245'."""
    if not code or code == 'ANOMALY':
        return None
    # Family = letter prefix + first digit
    letters = ''
    for c in code:
        if c.isalpha():
            letters += c
        else:
            break
    digits = code[len(letters):]
    if digits:
        return letters + digits[0]
    return letters


def draft_triage(aligned_rows):
    """
    Add draft divergence assessments to non-MATCH rows.

    Modifies rows in place, adding 'divergence_draft' field.
    """
    # Build indexes for cross-referencing
    gt_codes = set()
    matched_codes = set()
    ai_extra_codes = defaultdict(list)

    for r in aligned_rows:
        if r['type'] == 'MATCH':
            matched_codes.add(r['bullard_code'])
            gt_codes.add(r['bullard_code'])
        elif r['type'] == 'MISS':
            gt_codes.add(r['bullard_code'])

    for r in aligned_rows:
        if r['type'] == 'AI EXTRA':
            ai_extra_codes[r['ai_code']].append(r)

    # Build family sets
    gt_families = {get_family(c) for c in gt_codes if get_family(c)}
    ai_extra_families = defaultdict(list)
    for r in aligned_rows:
        if r['type'] == 'AI EXTRA':
            fam = get_family(r['ai_code'])
            if fam:
                ai_extra_families[fam].append(r)

    # Triage MISS rows
    for r in aligned_rows:
        if r['type'] == 'MISS':
            miss_family = get_family(r['bullard_code'])

            # Check if an AI EXTRA exists in the same family
            if miss_family and miss_family in ai_extra_families:
                candidates = [x['ai_code'] for x in ai_extra_families[miss_family]]
                r['divergence_draft'] = (
                    f"[DRAFT] REVIEW — possible family substitution. "
                    f"AI EXTRA candidates in same family: {', '.join(set(candidates))}"
                )
            # Check if AI used this code elsewhere
            elif r['bullard_code'] in [x['ai_code'] for x in aligned_rows if x['type'] == 'AI EXTRA']:
                r['divergence_draft'] = (
                    f"[DRAFT] REVIEW — AI assigned {r['bullard_code']} but on different text"
                )
            else:
                r['divergence_draft'] = "[DRAFT] REVIEW — AI did not assign"

        elif r['type'] == 'AI EXTRA':
            if r['ai_code'] == 'ANOMALY':
                r['divergence_draft'] = "[DRAFT] ANOMALY — AI flagged dictionary gap"
            elif r['ai_code'] in matched_codes:
                r['divergence_draft'] = "[DRAFT] REVIEW — possible duplicate of matched code"
            elif get_family(r['ai_code']) in gt_families:
                r['divergence_draft'] = (
                    f"[DRAFT] GENUINE ALTERNATE — same family as GT codes"
                )
            else:
                r['divergence_draft'] = "[DRAFT] REVIEW — AI added, Bullard did not assign"
        else:
            r['divergence_draft'] = ''

    return aligned_rows


# ── Spreadsheet Generation ──

def build_spreadsheet(aligned_rows, case_number, profile, output_path,
                      prev_match=None, prev_miss=None, prev_extra=None):
    """Build the formatted QA triage spreadsheet."""

    match_count = sum(1 for r in aligned_rows if r['type'] == 'MATCH')
    miss_count = sum(1 for r in aligned_rows if r['type'] == 'MISS')
    extra_count = sum(1 for r in aligned_rows if r['type'] == 'AI EXTRA')
    anomaly_count = sum(1 for r in aligned_rows if r.get('ai_code') == 'ANOMALY')
    gt_total = match_count + miss_count
    ai_total = match_count + extra_count

    # Styles
    GREEN = PatternFill('solid', fgColor='C6EFCE')
    ORANGE = PatternFill('solid', fgColor='FCE4D6')
    BLUE = PatternFill('solid', fgColor='D6E4F0')
    HEADER_FILL = PatternFill('solid', fgColor='4472C4')
    HEADER_FONT = Font(bold=True, color='FFFFFF', size=11, name='Arial')
    TITLE_FONT = Font(bold=True, size=14, name='Arial')
    SUBTITLE_FONT = Font(bold=True, size=11, name='Arial')
    NORMAL_FONT = Font(size=10, name='Arial')
    BOLD_FONT = Font(bold=True, size=10, name='Arial')
    DRAFT_FONT = Font(size=10, name='Arial', italic=True, color='7030A0')
    ALERT_FONT = Font(bold=True, size=11, name='Arial', color='CC0000')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap_align = Alignment(wrap_text=True, vertical='top')

    wb = Workbook()

    # ── Sheet 1: QA ──
    ws = wb.active
    ws.title = f"{case_number} Calibration QA"

    ws.merge_cells('A1:M1')
    ws['A1'] = f'Case {case_number} — Phase 2b Calibration Comparison ({profile} Profile)'
    ws['A1'].font = TITLE_FONT

    ws.merge_cells('A2:M2')
    ws['A2'] = (f'Bullard GT: {gt_total} codes    AI Output: {ai_total} codes    '
                f'MATCH: {match_count}    MISS: {miss_count}    '
                f'AI EXTRA: {extra_count}    ANOMALY: {anomaly_count}')
    ws['A2'].font = SUBTITLE_FONT

    ws.merge_cells('A3:M3')
    ws['A3'] = 'GREEN = match    ORANGE = Bullard coded, AI missed    BLUE = AI coded, Bullard did not'
    ws['A3'].font = SUBTITLE_FONT

    ws.merge_cells('A4:M4')
    ws['A4'] = (
        'Divergence Type options:\n'
        '  BULLARD ERROR — AI better fits the dictionary\n'
        '  GENUINE ALTERNATE — both codes defensible\n'
        '  AI ERROR — Bullard is correct\n'
        '  REVIEW — unclear, needs source text\n'
        '  ANOMALY — genuine gap in Bullard\'s dictionary'
    )
    ws['A4'].font = NORMAL_FONT
    ws['A4'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[4].height = 80

    ws.merge_cells('A5:M5')
    if prev_match is not None:
        ws['A5'] = (f'Previous: {prev_match} match / {prev_miss} miss / {prev_extra} AI extra    |    '
                    f'This run ({profile}): {match_count} match / {miss_count} miss / {extra_count} AI extra')
        ws['A5'].font = ALERT_FONT
    else:
        ws['A5'] = f'Profile: {profile}    |    {match_count} match / {miss_count} miss / {extra_count} AI extra'
        ws['A5'].font = SUBTITLE_FONT

    # Headers
    headers = [
        'GT Seq', 'Bullard Code', 'Bullard Description', 'Bullard Source Text',
        'AI Code', 'AI Description', 'Result',
        'AI Sentence Fragment', 'AI Justification',
        'Voice Speaker', 'Voice Content',
        'Divergence Type', 'Your Notes', 'AI Seq', 'Chunk'
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
        cell.border = thin_border
    ws.row_dimensions[6].height = 30

    widths = [8, 12, 35, 40, 12, 35, 12, 40, 45, 14, 14, 35, 30, 10, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data rows
    fill_map = {'MATCH': GREEN, 'MISS': ORANGE, 'AI EXTRA': BLUE}

    for row_idx, r in enumerate(aligned_rows, 7):
        fill = fill_map[r['type']]
        values = [
            r['gt_seq'] if r['gt_seq'] != '' else '',
            r['bullard_code'],
            r['bullard_desc'],
            r['bullard_citation'],
            r['ai_code'],
            r['ai_desc'],
            r['type'],
            r['ai_citation'],
            r['ai_reasoning'],
            r.get('voice_speaker', ''),
            r.get('voice_content_type', ''),
            r.get('divergence_draft', ''),
            '',  # Your Notes — for the human reviewer
            r['ai_seq'] if r['ai_seq'] != '' else '',
            r['ai_chunk'] if r['ai_chunk'] != '' else '',
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = NORMAL_FONT
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = wrap_align

        # Purple italic for draft divergence column (column 12)
        div_cell = ws.cell(row=row_idx, column=12)
        if div_cell.value and '[DRAFT]' in str(div_cell.value):
            div_cell.font = DRAFT_FONT

    ws.freeze_panes = 'A7'
    ws.auto_filter.ref = f'A6:O{6 + len(aligned_rows)}'

    # ── Sheet 2: Summary ──
    ws2 = wb.create_sheet('Summary')
    ws2['A1'] = f'Case {case_number} — {profile} QA Summary'
    ws2['A1'].font = TITLE_FONT
    ws2.merge_cells('A1:C1')

    match_rate = f'{match_count / gt_total * 100:.1f}%' if gt_total > 0 else 'N/A'
    precision = f'{match_count / ai_total * 100:.1f}%' if ai_total > 0 else 'N/A'

    summary_data = [
        ('', '', ''),
        ('Metric', 'Previous', f'{profile} (current)'),
        ('Total GT codes', prev_match + prev_miss if prev_match is not None else 'N/A', gt_total),
        ('Total AI codes', prev_match + prev_extra if prev_match is not None else 'N/A', ai_total),
        ('MATCH', prev_match if prev_match is not None else 'N/A', match_count),
        ('MISS', prev_miss if prev_miss is not None else 'N/A', miss_count),
        ('AI EXTRA', prev_extra if prev_extra is not None else 'N/A', extra_count),
        ('ANOMALY', 'N/A', anomaly_count),
        ('Match Rate (of GT)',
         f'{prev_match / (prev_match + prev_miss) * 100:.1f}%' if prev_match is not None else 'N/A',
         match_rate),
        ('Precision (match/AI total)',
         f'{prev_match / (prev_match + prev_extra) * 100:.1f}%' if prev_match is not None else 'N/A',
         precision),
    ]

    for r_idx, (a, b, c) in enumerate(summary_data, 2):
        ws2.cell(row=r_idx, column=1, value=a).font = BOLD_FONT if r_idx == 3 else NORMAL_FONT
        ws2.cell(row=r_idx, column=2, value=b).font = BOLD_FONT if r_idx == 3 else NORMAL_FONT
        ws2.cell(row=r_idx, column=3, value=c).font = BOLD_FONT if r_idx == 3 else NORMAL_FONT
        for col in range(1, 4):
            ws2.cell(row=r_idx, column=col).border = thin_border

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 22

    # ── Sheet 3: Miss Analysis ──
    ws3 = wb.create_sheet('Miss Analysis')
    ws3['A1'] = f'Case {case_number} — Missed Codes Breakdown'
    ws3['A1'].font = TITLE_FONT
    ws3.merge_cells('A1:E1')

    miss_headers = ['GT Seq', 'Bullard Code', 'Bullard Description', 'Family', 'Draft Assessment']
    for col_idx, h in enumerate(miss_headers, 1):
        cell = ws3.cell(row=3, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = thin_border

    miss_row = 4
    family_counts = defaultdict(int)
    for r in aligned_rows:
        if r['type'] == 'MISS':
            fam = get_family(r['bullard_code']) or '?'
            family_counts[fam] += 1
            vals = [r['gt_seq'], r['bullard_code'], r['bullard_desc'], fam,
                    r.get('divergence_draft', '')]
            for col_idx, v in enumerate(vals, 1):
                cell = ws3.cell(row=miss_row, column=col_idx, value=v)
                cell.font = NORMAL_FONT
                cell.fill = ORANGE
                cell.border = thin_border
                cell.alignment = wrap_align
            miss_row += 1

    # Family summary
    miss_row += 1
    ws3.cell(row=miss_row, column=1, value='Family Distribution').font = BOLD_FONT
    miss_row += 1
    for fam, count in sorted(family_counts.items(), key=lambda x: -x[1]):
        ws3.cell(row=miss_row, column=1, value=fam).font = NORMAL_FONT
        ws3.cell(row=miss_row, column=2, value=count).font = NORMAL_FONT
        miss_row += 1

    ws3.column_dimensions['A'].width = 10
    ws3.column_dimensions['B'].width = 14
    ws3.column_dimensions['C'].width = 45
    ws3.column_dimensions['D'].width = 10
    ws3.column_dimensions['E'].width = 50

    wb.save(output_path)
    return {
        'match': match_count, 'miss': miss_count, 'extra': extra_count,
        'anomaly': anomaly_count, 'gt_total': gt_total, 'ai_total': ai_total,
        'match_rate': match_rate, 'precision': precision,
        'family_counts': dict(family_counts),
    }


# ── CLI ──

def main():
    parser = argparse.ArgumentParser(description='Phase 2b QA Triage Generator')
    parser.add_argument('--input', '--csv', dest='input_path', required=True,
                        help='Path to phase2 results (CSV or JSON, auto-detected)')
    parser.add_argument('--db', required=True, help='Path to ufo_matrix.db')
    parser.add_argument('--case', required=True, help='Case number (e.g., 192g)')
    parser.add_argument('--profile', default='unknown', help='Prompt profile name')
    parser.add_argument('--output', required=True, help='Output .xlsx path')
    parser.add_argument('--prev-match', type=int, default=None, help='Previous MATCH count')
    parser.add_argument('--prev-miss', type=int, default=None, help='Previous MISS count')
    parser.add_argument('--prev-extra', type=int, default=None, help='Previous AI EXTRA count')

    args = parser.parse_args()

    print(f"Loading AI results from {args.input_path}...")
    ai_rows = load_ai_results(args.input_path)
    print(f"  {len(ai_rows)} AI-coded events")

    print(f"Loading ground truth for case {args.case} from {args.db}...")
    gt_rows = load_ground_truth(args.db, args.case)
    print(f"  {len(gt_rows)} ground truth events")

    print("Loading motif dictionary...")
    motif_dict = load_motif_dict(args.db)
    print(f"  {len(motif_dict)} motif codes")

    print("Running alignment...")
    aligned = align(gt_rows, ai_rows, motif_dict)

    print("Running draft triage...")
    aligned = draft_triage(aligned)

    print(f"Building spreadsheet: {args.output}")
    stats = build_spreadsheet(
        aligned, args.case, args.profile, args.output,
        prev_match=args.prev_match, prev_miss=args.prev_miss, prev_extra=args.prev_extra
    )

    print(f"\n{'='*50}")
    print(f"  MATCH:     {stats['match']}")
    print(f"  MISS:      {stats['miss']}")
    print(f"  AI EXTRA:  {stats['extra']}")
    print(f"  ANOMALY:   {stats['anomaly']}")
    print(f"  Match Rate: {stats['match_rate']}")
    print(f"  Precision:  {stats['precision']}")
    if stats['family_counts']:
        print(f"\n  Miss families: {stats['family_counts']}")
    print(f"{'='*50}")
    print(f"\nSaved to: {args.output}")

    # Output JSON stats for programmatic use
    print(f"\nSTATS_JSON:{json.dumps(stats)}")


if __name__ == '__main__':
    main()
